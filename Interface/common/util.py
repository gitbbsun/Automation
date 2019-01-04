#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/12/28 17:01
# @Author  : bbsun
# @File    : util.py
# @Software: PyCharm

import json, xlrd, requests
import openpyxl
from openpyxl import load_workbook

base_url = "https://pre.svocloud.com/"


# xlrd读取excel测试数据，返回字典格式
class ReadExcel():
    def __init__(self, excelPath, sheetName="Sheet1"):
        self.data = xlrd.open_workbook(excelPath)
        self.table = self.data.sheet_by_name(sheetName)
        # 获取第一行作为总行数
        self.keys = self.table.row_values(0)
        # 获取总行数
        self.rowNum = self.table.nrows
        # 获取总列数
        self.colNum = self.table.ncols

    def dict_data(self):
        if self.rowNum <= 1:
            print("表格未填数据")
            return None
        else:
            # 获取表格第一行列名
            keys = self.table.row_values(0)
            # print(keys)
            ExcelData = []
            # 获取每一行的内容，列表格式
            for col in range(1, self.rowNum):
                values = self.table.row_values(col)
                # keys,values这2个列表一一对应来组合转换为字典
                api_data = dict(zip(keys, values))
                ExcelData.append(api_data)
            return ExcelData


def send_requests(case_data):
    method = case_data["method"]
    url = base_url + case_data["url"]
    headers = case_data["headers"]
    body = case_data["body"]
    params = case_data["params"]
    if headers == "":
        headers = {}
    else:
        headers = json.loads(case_data["headers"])

    if params == "":
        params = {}
    else:
        params = json.loads(case_data["params"])
    if body == "":
        body = {}
    else:
        body = json.loads(case_data["body"])

    if (method == "get"):
        re = requests.post(url, headers=headers, params=params)
    else:
        re = requests.post(url, json=body, headers=headers)
    return re.status_code

    # verify = False
    # res = {}

    # if method == "post":
    #     try:
    #         r = s.request(method=method,
    #                       url=url,
    #                       params=params,
    #                       headers=headers,
    #                       data=body,
    #                       verify=verify)
    #         res['id'] = testdata['id']
    #         res['rowNum'] = testdata['rowNum']
    #         res['statuscode'] = str(r.status_code)
    #         res['text'] = r.content.decode("utf-8")
    #         res["times"] = str(r.elapsed.total_seconds())
    #         if res["statuscode"] != "200":
    #             res["error"] = res["text"]
    #         else:
    #             res["error"] = ""
    #         res["msg"] = ""
    #         if testdata["checkpoint"] in res["text"]:
    #             res["result"] = "pass"
    #         else:
    #             res["result"] = "fail"
    #         return res
    #     except Exception as msg:
    #         res["msg"] = str(msg)
    #         return res


def copy_excel(data_path, report_path):
    report = openpyxl.Workbook()
    report.save(report_path)
    # 读取数据
    data = openpyxl.load_workbook(data_path)
    report = openpyxl.load_workbook(report_path)
    sheets1 = data.sheetnames
    sheets2 = report.sheetnames
    sheet1 = data[sheets1[0]]
    sheet2 = report[sheets2[0]]
    max_row = sheet1.max_row
    max_column = sheet1.max_column
    for m in list(range(1, max_row + 1)):
        for n in list(range(97, 97 + max_column)):  # chr(97)=='a'
            n = chr(n)
            i = f"{n}{m}"  # 单元格编号
            cell1 = sheet1[i].value  # 获取data单元格数据
            sheet2[i].value = cell1  # 赋值到test单元格
    report.save(report_path)
    data.close()
    report.close()


class Write_excel(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet

    def write(self, row_n, col_m, value):
        self.ws.cell(row_n, col_m).value = value
        self.wb.save(self.filename)
